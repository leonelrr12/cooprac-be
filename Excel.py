
import mysql.connector

def insert_acp(id, titulo, grupo, segmento, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO profesions_acp (id,titulo,grupo,segmento) VALUES (%s,%s,%s,%s)"

        record = (id,titulo,grupo,segmento)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            # cnn.close()
            print("MySQL connection is closed")


def insert_lw(id, titulo, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO profesions_lw (id,titulo) VALUES (%s,%s)"

        record = (id,titulo)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            # cnn.close()
            print("MySQL connection is closed")


def insert_inst(id, name, subgrupo, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO institutions (id,name,subgrupo) VALUES (%s,%s,%s)"

        record = (id,name,subgrupo)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            # cnn.close()
            print("MySQL connection is closed")


def insert_plan(id, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO planillas_j (id,name) VALUES (%s,%s)"

        record = (id,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_pais(id, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO nationality (id,name,is_active) VALUES (%s,%s,1)"

        record = (id,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_prov(id, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO provinces (id,name) VALUES (%s,%s)"

        record = (id,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_dist(id, idProv, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO districts (id, idProv, name) VALUES (%s,%s,%s)"

        record = (id,idProv,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


def insert_corr(id, idDist, idProv, name, cnn):
    try:
        cursor = cnn.cursor()
        mySql_insert_query = "INSERT INTO counties (id,  idProv, idDist, name) VALUES (%s,%s,%s,%s)"

        record = (id,idDist,idProv,name)
        cursor.execute(mySql_insert_query, record)
        # cnn.commit()
        # print("Record inserted!")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
    
    finally:
        if cnn.is_connected():
            cursor.close()
            print("MySQL connection is closed")


# # librerias para correos
# import smtplib, ssl
# import getpass
# libresia para excel
import openpyxl

def fcnn():
    return mysql.connector.connect(
    host="finanservs.cdxgqqkffqvx.us-east-2.rds.amazonaws.com",
    user="admin",
    password="Cooprac2616",
    database="finanservs",
    port="3306"
)

# # No. 1
def acp():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\Profesiones_ACP.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'D751']

    # MySql de Digital Ocean
    cnn = fcnn()

    print("Profsionales ACP ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_acp(data[0], data[1], data[2], data[3], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

# # No. 2
def lw():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\Profesiones_Linea_Blanca.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A5' : 'B105']
    cnn = fcnn()

    print("Linea Blanca ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_lw(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

# No. 3
def inst():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\Instituciones.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A3' : 'C70']

    cnn = fcnn()

    print("Instituciones ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_inst(data[0], data[1], data[2], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

# # No. 4
def plan():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\Planillas_Jubilados.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A3' : 'B15']

    cnn = fcnn()

    print("Planillas Jubilados ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_plan(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

# # No. 567
def pais():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\paises_estandar.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'B238']

    cnn = fcnn()

    print("Paises ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_pais(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")


# # No. 5
def prov():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\prov-codigo.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'B11']

    cnn = fcnn()

    print("Provincias ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_prov(data[0], data[1], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

# # No. 6
def dist():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\dist-codigo.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'C71']

    cnn = fcnn()

    print("Distritos ...") 
    for fila in celdas:
        data = [celda.value for celda in fila]
        # print(data)
        insert_dist(data[0], data[1], data[2], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")

# # No. 7
def corr():
    fichero = r"D:\Documentos\Desarrollo Web\Finanservs\corr-codigo.xlsx"

    book = openpyxl.load_workbook(fichero, data_only=True)
    hoja = book.active

    celdas = hoja['A2' : 'D599']

    cnn = fcnn()

    print("Corregimientos ...")
    for fila in celdas:
        data = [celda.value for celda in fila]
        insert_corr(data[0], data[1], data[2], data[3], cnn)

    cnn.commit()
    if cnn.is_connected():
        cnn.close()
        print("MySQL Finish ...")



#acp()
#lw()
#inst()
#plan()
pais()
#prov()
#dist()
#corr()
